import os
from pprint import pprint
import openpyxl

class ExcelHandler:

    def __init__(self,file_path,sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def excel_test(self):

        self.workbook = openpyxl.load_workbook(self.file_path)
        workbook = self.workbook[self.sheet_name]
        rows = list(workbook.rows)
        headers = []
        for title in rows[0]:
            headers.append(title.value)
        data = []
        for row in rows[1:]:
            read_data={}
            for idx, cell in enumerate(row):
                read_data[headers[idx]] = cell.value
            data.append(read_data)
        self.close()
        return data

    def write(self, sheet_name, row, column, data):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row, column).value= data
        self.save()
        self.close()

    def save(self):
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()

if  __name__ == '__main__':

    excel = ExcelHandler("cases.xlsx", "add")
    excel = excel.excel_test()
    pprint(excel)

























