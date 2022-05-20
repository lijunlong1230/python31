import os.path
from pprint import pprint

from jsonpath import jsonpath
from openpyxl.worksheet.worksheet import Worksheet

#
# def test1():
#     lst = []
#     for i in range(3):
#         n = int(input("输入三个整数："))
#         lst.append(n)
#     lst.sort()
#     print(lst)
#
# def test2(n):
#     r = 1
#     for i in range(2, n+1):
#         r *= i
#         print(r)
#
# def test3(n):
#     a,b = 1,1
#     for i in range(n-1):
#         a,b = b,a+b
#     return b
# test2(5)
# print(test3(5))


# score = int(input("请输入成绩："))
# if score>=90:
#     print("A")
# elif 60<=score<=90:
#     print("B")
# else:
#     score<60
#     print("C")

# #'2+22+222+2222+22222'
# a = int(input("循环几次："))
# b = str(input("要加的数字："))
# n = ""
# m = ""
# for i in range(a):
#     n += b
#     m += '+' + n
#
# pprint(m[1:])

#!/usr/bin/python # -*- coding: UTF-8 -*-
# tour = []
# height = []
# hei = 100.0
# # 起始高度
# tim = 10
# # 次数
# for i in range(1, tim + 1):
#     tour.append(hei)
#     hei /= 2
#     height.append(hei)
# print('总高度：tour = {0}'.format(sum(tour)))
# print('第10次反弹高度：height = {0}'.format(height[-1]))
#
#
# #猴子吃桃没天吃一个零一个最后剩一个 一共多少桃子
# x = 1
# for i in range(9,0,-1):
#     x = (x+1)*2
#     print(x)
# print(x)


# def up(n):#定义菱形的上半部分
#     for line in range(n):#行数
#         for space_count in range(n-line-1):#打印每行前面的空格，line加一则空格减一
#             print(" ",end="")#print默认是打印一行，结尾加换行。end=' '意思是末尾不换行，加空格
#         for start in range(line+1):#打印星星，*随着line增加而增加
#             print("* ",end='')
#         print()
#
#
# def down(n):#菱形的下半部分3
#
#     for line in range(n-1):#从菱形的中间下一行打印
#         for space_count in range(line+1):
#             print(" ",end="")
#         for start in range(n-line-1):
#             print("* ",end='')
#         print()
# n=int(input("请输入等边菱形的边长: "))
# up(n)
# down(n)

# #
# n = int(input("qingshuru***jige:"))
# for i in range(n):
#     for j in range(n-i-1):
#         print(" ",end="")#end=' '意思是末尾不换行，加空格
#     for k in range(i+1):
#         print("* ",end="")
#     print()
# for j in range(n-1):
#     for i in range(j+1):
#         print(" ",end="")
#     for k in range(n-j-1):
#         print("* ",end="")
#     print()
#
# import re
# mystring = 'jlkjadfjs2:13409rf029f@O)(UIR)(#FOIj2f'
# print(re.search(r".+?", mystring))
# print(re.match(r".*?", mystring))
# print(re.findall(r"[a-z]", mystring))
# print(re.sub(r"\d", "909", mystring, count=2))
#
# print("============")
# print(re.search(r".", mystring))

import openpyxl

# #第一步打开文件
# workbook = openpyxl.load_workbook('cases.xlsx')
# #第二部获取表单
# sheet: Worksheet = workbook['login']
#
# #第三部 获取某一行某一列单元格
# cell = sheet.cell(row=3, column=4)
#
# rows = list(sheet.rows)
# print(rows)
# for row in rows[:]:
#     print(type(row))
#     for cell in row:
#         print(type(cell.value))

import openpyxl

workbook = openpyxl.load_workbook("cases.xlsx")
print(workbook)
sheet: Worksheet = workbook["add"]
rows = list(sheet.rows)

for i in rows[1:]:
    print(i)
    for j in i:
        print(j.value)












